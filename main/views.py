from django.shortcuts import render
#from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.views import generic
from django.template.loader import get_template
from django .utils import timezone
from xhtml2pdf import pisa
from main.models import *
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView,ModelFormMixin,DeleteView
from django.views.generic.detail import DetailView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.urls import reverse_lazy
from main.forms import CandidatoCreateForm
#,ActaCreateForm
from django.http import HttpResponse
from django.core import serializers
from django.db.models import Count, Avg, Func, F, Case,When, CharField, Sum, Value
import json
from main.forms import *
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
###Dashboard
from django.http import JsonResponse
from django.shortcuts import render
#from dashboard.models import Order
from django.core import serializers
#
# Create your views here.
#class home(LoginRequiredMixin,generic.TemplateView):

# Utility function

def link_callback(uri, rel):
            """
            Convert HTML URIs to absolute system paths so xhtml2pdf can access those
            resources
            """
            result = finders.find(uri)
            if result:
                    if not isinstance(result, (list, tuple)):
                            result = [result]
                    result = list(os.path.realpath(path) for path in result)
                    path=result[0]
            else:
                    sUrl = settings.STATIC_URL        # Typically /static/
                    sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
                    mUrl = settings.MEDIA_URL         # Typically /media/
                    mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

                    if uri.startswith(mUrl):
                            path = os.path.join(mRoot, uri.replace(mUrl, ""))
                    elif uri.startswith(sUrl):
                            path = os.path.join(sRoot, uri.replace(sUrl, ""))
                    else:
                            return uri

            # make sure that file exists
            if not os.path.isfile(path):
                    raise Exception(
                            'media URI must start with %s or %s' % (sUrl, mUrl)
                    )
            return path


class home(LoginRequiredMixin, generic.TemplateView):
    template_name="base/home.html"

    def get_context_data(self,*args, **kwargs):
      context = super(home, self).get_context_data(**kwargs)

      return context

class estadistica( generic.TemplateView):
    template_name="estadisticas/estadisticas.html"

    def get_context_data(self,*args, **kwargs):
      context = super(estadistica, self).get_context_data(**kwargs)
      #documentos = TblDocument.objects.filter(active = 1).annotate(cantidad=Count('id_document'))
      #documentos = TblDocument.objects.all().count()
      #cargados = TblDocument.objects.filter(active = 1).count()
      #enproceso = TblDocument.objects.filter(active = 2, documento_acta__isnull = True).count()
      #ingresadas = TblDocument.objects.filter(active = 3).count()
      departamentos = Departamento.objects.all()
      municipios = Municipio.objects.all()
      tipos = Tipo.objects.all()

      m1 = Escrutinio.objects.filter(candidato__tipo = 1).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
      presidencial = Escrutinio.objects.filter(candidato__tipo = 1).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('num_votos')


      context['escrutinio'] = m1
      #context['documentos'] = documentos
      #context['enproceso'] = enproceso
      #context['cargados'] = cargados
      #context['ingresadas'] = ingresadas
      context['departamentos'] = departamentos
      context['municipios'] = None
      #context['tipos'] = tipos
      context['error'] = 0
      context['descripcion'] = 'Nivel Nacional'
      context['presidencial']= presidencial

      return context


def filtrar(request):
    print(request)
    if request.method == 'GET':
        error = -1
        partidos_d = None

        departamento = int(request.GET.get('departamento'))
        print(departamento)
        if request.GET.get('municipio'):
            municipio = int(request.GET.get('municipio'))
        else:
            municipio = None

        print(departamento)
        #Presidencial

        if departamento == 0:
            print('Nivel Nacional')
            conteo_centros = CentroEducativo.objects.count()
            conteo_mesa = Mesa.objects.count()
            conteo_mesa_registrada = Escrutinio.objects.values('mesa_id').distinct().count()
            presidencial = Escrutinio.objects.filter(candidato__tipo = 1).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')

            porc = round((conteo_mesa_registrada/conteo_mesa)*100,2)
            diputaciones = None
            municipales = None
            municipio = 0
            municipios = -1
            descripcion = 'Nivel Nacional'
        #Sí es Departamental
        if departamento > 0:
            conteo_mesa_registrada = Escrutinio.objects.filter(mesa__centro__departamento__pk = departamento).values('mesa_id').distinct().count()
            conteo_mesa = Mesa.objects.filter(centro__departamento__pk= departamento).count()
            conteo_centros = CentroEducativo.objects.filter(departamento = departamento).count()
            partidos_d = Escrutinio.objects.filter(candidato__tipo = 3).values('candidato__partido','candidato__partido__nombre','candidato__partido__logo').distinct()
            print('Departamental')
            descripcion = 'Nivel Departamental'
            #escrutinio = Escrutinio.objects.filter(candidato__tipo = 2 ).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('num_votos')
            diputaciones = Escrutinio.objects.filter(candidato__tipo = 3,mesa__centro__departamento__pk = departamento).values('candidato__partido','candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            municipales = None
            #Escrutinio.objects.filter(candidato__tipo = 2,mesa__centro__departamento__pk = departamento).values('candidato__partido','candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            presidencial = Escrutinio.objects.filter(candidato__tipo = 1,mesa__centro__departamento__pk = departamento).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            print(presidencial.query)
            porc = round((conteo_mesa_registrada/conteo_mesa)*100,2)
            municipios = Municipio.objects.all()

        if request.GET.get('municipio') and municipio > 0:
            print('Departamental')
            descripcion = 'Nivel Departamental'
            conteo_mesa = Mesa.objects.filter(centro__municipio__pk= municipio).count()
            conteo_mesa_registrada = Escrutinio.objects.filter(mesa__centro__municipio__pk = municipio).values('mesa_id').distinct().count()
            conteo_centros = CentroEducativo.objects.filter(municipio = municipio).count()
            partidos_d = Escrutinio.objects.filter(candidato__tipo = 3).values('candidato__partido','candidato__partido__nombre','candidato__partido__logo').distinct()
            municipales = Escrutinio.objects.filter(candidato__tipo = 2,mesa__centro__municipio__pk = municipio).values('candidato__partido','candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            diputaciones = Escrutinio.objects.filter(candidato__tipo = 3,mesa__centro__municipio__pk = municipio).values('candidato__partido','candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            escrutinio = Escrutinio.objects.filter(candidato__tipo = 2 ).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            presidencial = Escrutinio.objects.filter(candidato__tipo = 1,mesa__centro__municipio__pk = municipio).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('-num_votos')
            print(presidencial.query)
            porc = round((conteo_mesa_registrada/conteo_mesa)*100,2)
            municipios = Municipio.objects.all()


        departamentos = Departamento.objects.all()
        tipos = Tipo.objects.all()

        #identidad_ = request.GET.get('identidad')

        return render(request,"estadisticas/estadisticas.html",{'municipio':municipio,'departamento':departamento,'departamentos':departamentos, 'municipios':municipios,'descripcion':descripcion, 'presidencial':presidencial,'diputaciones':diputaciones, 'error':error, 'partidos': partidos_d, 'conteo_centros': conteo_centros, 'conteo_mesas': conteo_mesa, 'conteo_mesa_registrada': conteo_mesa_registrada,'porc':porc,'municipales':municipales},)

class EscrutinioList(ListView):
    model = Mesa
    template_name= 'main/escrutinio_list.html'
    #paginate_by = 30
    context_object_name = "obj"
        #print('Listar productos')
    def get_queryset(self):
        p = Mesa.objects.all()
        print(p)
        return p

class EscrutinioCreate(SuccessMessageMixin,CreateView):
    model = Escrutinio
    form = EscrutinioCreateForm
    fields = [
                'mesa',
                'candidato',
                'valor',
            ]
    success_url= reverse_lazy("main:escrutinio-list")
    success_message="Escrutinio creada satisfactoriamente"
    error_message="Error, favor revise los campos ya que el acta no se puede crear"

    def post(self, request, *args, **kwargs):
        print('Haciendo Post')
        # Load form
        form = self.get_form()
        print(form)

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super(EscrutinioCreate, self).get_context_data(**kwargs)
        candidatos = None
        centros_votacion = Escrutinio.objects.values('mesa').distinct()
        centros = Mesa.objects.exclude(pk__in=centros_votacion)
        context['candidatos'] = candidatos
        context['centros'] = centros

        return context

class MesaCreate(SuccessMessageMixin,CreateView):
    model = Mesa
    fields = [
            'centro',
            'descripcion',
            'numero',
        ]

    success_url= reverse_lazy("main:mesa-add")
    success_message="Mesa Creada Satisfactoriamente"
    error_message="Error, favor revise los campos ya que la mesa no se puede crear"


    def post(self, request, *args, **kwargs):
        form = MesaCreateForm(request.POST, request.FILES)
        print(request.POST)
        id_numero = request.POST['numero']
        print('-------------')
        print(id_numero)
        exito = 0
        error = 0
        guardar = True
        if  len(Mesa.objects.filter(numero = id_numero)) > 0:
            guardar = False
            error = 2

        if form.is_valid() and guardar:
            print('Formulario valido')
            try:
                mesa = form.save(commit=False)
                mesa.user_created_id = self.request.user.pk
                mesa.user_update_id = self.request.user.pk
                mesa.save()
                form.save()
                exito = 1
                print('exito')
                print(exito)
                return render(request, 'main/mesa_form.html', {'form': form, 'exito': exito,'obj': obj})
            except Exception as e:
                error = 1
                return render(request, 'main/mesa_form.html', {'form': form, 'error': error})
        else:
            if guardar:
                error = 1
            else:
                error = 2

            centro = CentroEducativo.objects.all()
            #municipio = Municipio.objects.all()

            print('Formulario Erroneo')
            print(error)
            return render(request, 'main/mesa_form.html', {'form': form, 'error': error})


class CandidatoCreate(SuccessMessageMixin,CreateView):
    model = Candidato
    fields = [
                'tipo',
                'partido',
                'identidad',
                'nombre',
                'Sexo',
                'foto',
                'departamento',
                'municipio'
            ]
    success_url= reverse_lazy("main:candidato-add")
    success_message="Candidato Creado Satisfactoriamente"
    error_message="Error, favor revise los campos ya que el candidato no se puede crear"

    def post(self, request, *args, **kwargs):
        form = CandidatoCreateForm(request.POST, request.FILES)
        exito = 1
        error = 1
        departamento = Departamento.objects.all()
        municipio = Municipio.objects.all()
        obj = Candidato.objects.all()

        if form.is_valid():
            print('Formulario valido')

            try:
                candidato = form.save(commit=False)
                candidato.user_created_id = self.request.user.pk
                candidato.user_update_id = self.request.user.pk
                candidato.save()
                form.save_m2m()

                return render(request, 'main/candidato_form.html', {'form': form, 'exito': exito,'obj': obj, 'departamentos': departamento, 'municipios': municipio })
            except Exception as e:
                return render(request, 'main/candidato_form.html', {'form': form, 'error': error, 'departamentos': departamento, 'municipios': municipio})
        else:
            departamento = Departamento.objects.all()
            municipio = Municipio.objects.all()

            print('Formulario Erroneo')
            return render(request, 'main/candidato_form.html', {'form': form, 'error': error, 'departamentos': departamento, 'municipios': municipio})

    def get_context_data(self, **kwargs):
        context = super(CandidatoCreate, self).get_context_data(**kwargs)
        #slug = self.kwargs['pk']
        departamento = Departamento.objects.all()
        municipio = Municipio.objects.all()
        #print(slug)
        context['departamentos'] = departamento
        context['municipios'] =  municipio
        return context


class CandidatoUpdate(SuccessMessageMixin,UpdateView,ModelFormMixin):
    model = Candidato
    fields = [
                'tipo',
                'partido',
                'identidad',
                'nombre',
                'Sexo',
                #'telefono',
                #'email',
                #'direccion',
                'foto',
            ]
    success_url= reverse_lazy("candidato-add")
    success_message="Candidato Modificado Satisfactoriamente"
    error_message="Error, favor revise los campos ya que el candidato no se puede crear"

    def get_context_data(self, **kwargs):
        context = super(CandidatoUpdate, self).get_context_data(**kwargs)
        #slug = self.kwargs['pk']
        departamento = Departamento.objects.all()
        municipio = Municipio.objects.all()
        #print(slug)
        context['departamentos'] = departamento
        context['municipios'] =  municipio
        return context

    def post(self, request, *args, **kwargs):
        #print('En el post')
        # Get instance of PhysicalPart
        self.object = self.get_object()
        departamento = Departamento.objects.all()
        municipio = Municipio.objects.all()
        # Load form
        form = self.get_form()
        # Add choices to form 'subcategory' field
        #form.fields['subcategory'].choices = SubcategoryFilter[self.object.category]
        #print(form)
        # Check if form is valid and save PhysicalPart instance
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

class CandidatoDelete(DeleteView):
    model = Candidato
    success_url= reverse_lazy("candidato-list")

class CandidatoList(ListView):
    model = Candidato
    #paginate_by = 30
    context_object_name = "obj"
        #print('Listar productos')
    def get_queryset(self):
        p = Candidato.objects.all()
        return p

# class TblDocumentList(ListView):
#     template_name="main/tblodocument_list.html"
#     model = TblDocument
#     #paginate_by = 30
#     context_object_name = "obj"
#         #print('Listar productos')
#     def get_queryset(self):
#
#         p = TblDocument.objects.filter(active = 1)
#         print(p.query)
#         return p
#
# class ListProceso(generic.TemplateView):
#     template_name="main/list_proceso.html"
#     #model = TblDocument
#     #paginate_by = 30
#     context_object_name = "obj"
#         #print('Listar productos')
#     def get_context_data(self,*args, **kwargs):
#       context = super(ListProceso, self).get_context_data(**kwargs)
#       obj = TblDocument.objects.filter(active = 2)
#       context['obj'] = obj
#       return context
#
#     def get_queryset(self):
#         obj = TblDocument.objects.filter(active = 2)
#         return obj


class MesaList(generic.TemplateView):
    template_name="main/Mesa_list.html"
    #model = TblDocument
    #paginate_by = 30
    context_object_name = "obj"
        #print('Listar productos')
    def get_context_data(self,*args, **kwargs):
      context = super(MesaList, self).get_context_data(**kwargs)
      obj = Mesa.objects.all()
      context['obj'] = obj
      return context

    def get_queryset(self):
        obj = Mesa.objects.all()
        return obj


def guardar_acta_escrutinio(request):
    print(request.POST)
    json_obj = json.loads(request.POST['votos'])
    mesa = request.POST['centro_educativo']

    centro_ = Mesa.objects.get(pk = mesa)

    print(mesa)
    #print(docu)
    for value in json_obj:
            candidato = Candidato.objects.get(pk = int(value['id']))
            try:
                obj1, created = Escrutinio.objects.update_or_create(
                                    mesa = centro_,
                                    candidato = candidato,
                                    defaults = {
                                        'candidato': candidato,
                                        #'corriente': corriente,
                                        'user_created':request.user,
                                        'user_update':request.user,
                                        'valor': int(value['new_weight'])
                                        }
                                        )
                message = 1
                print(message)
            except Exception as e:
                print('Error en votos /*/*/*/*/*/' , e)
                message = 0

    print(message)
    #print(json_obj2)
    #for value in json_obj:
    #    print(value['id'])
    #temp_arr.append(value[0]["firstName"])
    return HttpResponse(message)

def cambiar_estado(request):
    #print(request.POST)
    acta = request.POST['acta']
    print('El acta -- -- -- ' , acta)
    message = 3
    acta_ = TblDocument.objects.get(id_document__in = acta)
    try:
        acta_.active = 2
        acta_.save()

        message = 1
    except Exception as e:
            print('Error -----' , e)
            message = 0
    return HttpResponse(message)


class EscrutinioDetailView(DetailView):
    model = Mesa

    def get_queryset(self):
        print('Modificando el queryset')
        slug = self.kwargs['pk']
        p = Mesa.objects.filter(pk = slug)
        print(p)
        return p

    def get_context_data(self, **kwargs):
        #print('en la vista')
        #obtenemos el numero de centro educativo
        slug = self.kwargs['pk']
        #print(' - - - - - - -')
        #print(slug)

        context = super().get_context_data(**kwargs)
        context['now'] = 'hoy'
        #documento = Escrutinio.objects.all(acta__documento_id = self.request.user.pk)[:1]
        obj = Escrutinio.objects.filter(mesa__pk = int(slug)).order_by('candidato__partido','candidato__tipo')
        #print(obj.query)
        centro = Mesa.objects.filter(pk = slug)
        #print(obj)
        context['obj'] = obj
        context['centro'] = centro

        return context

class ColaboradorCreate(SuccessMessageMixin,CreateView):
    model = Colaborador
    form = ColaboradorCreateForm
    fields = [
        'municipio',
        'user',
        'tipo',
        'sexo',
        'identidad',
        'nombre',
        'foto'

    ]
    success_url= reverse_lazy("main:colaborador-add")
    success_message="Colaborador agregado satisfactoriamente"
    error_message="Error, favor revise los campos ya que el colaborador no se puede crear"

    def post(self, request, *args, **kwargs):
        form = ColaboradorCreateForm(request.POST, request.FILES)
        print(request.POST['user'])
        user = request.POST['user']
        usuario = User.objects.filter(pk= user)
        exito = 0
        error = 0
        obj = Colaborador.objects.all()

        if usuario:
            error = 3
            return render(request, 'main/colaborador_form.html', {'form': form, 'error': error,})
        else:
            if form.is_valid():
                print('Formulario valido')
                try:
                    colaborador = form.save(commit=False)
                    colaborador.user_created_id = self.request.user.pk
                    colaborador.user_update_id = self.request.user.pk
                    colaborador.save()
                    #form.save_m2m()
                    return render(request, 'main/colaborador_form.html', {'form': form, 'exito': exito,'obj': obj})
                except Exception as e:
                    return render(request, 'main/colaborador_form.html', {'form': form, 'error': error,})
            else:
                print('Formulario Erroneo')
                return render(request, 'main/colaborador_form.html', {'form': form, 'error': error})



class ColaboradorUpdate(SuccessMessageMixin,UpdateView,ModelFormMixin):
    model = Colaborador
    fields = [
                'municipio',
                'user',
                'tipo',
                'sexo',
                'identidad',
                'nombre',
                'foto'
            ]
    success_url= reverse_lazy("colaborador-add")
    success_message="Colaborador Modificado Satisfactoriamente"
    error_message="Error, favor revise los campos ya que el colaborador no se puede actualizar "

    # def get_context_data(self, **kwargs):
    #     context = super(CandidatoUpdate, self).get_context_data(**kwargs)
    #     #slug = self.kwargs['pk']
    #     departamento = Departamento.objects.all()
    #     municipio = Municipio.objects.all()
    #     #print(slug)
    #     context['departamentos'] = departamento
    #     context['municipios'] =  municipio
    #     return context


    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        # Load form
        form = self.get_form()
        user = request.POST['user']
        usuario = User.objects.filter(pk= user)
        exito = 0
        error = 0
        obj = Colaborador.objects.all()

        if form.is_valid():
            print('Formulario valido')
            try:
                colaborador = form.save(commit=False)
                colaborador.user_created_id = self.request.user.pk
                colaborador.user_update_id = self.request.user.pk
                colaborador.save()
                #form.save()
                exito=1
                return render(request, 'main/colaborador_form.html', {'form': form, 'exito': exito,'obj': obj})
            except Exception as e:
                error=1
                print(e)
                return render(request, 'main/colaborador_form.html', {'form': form, 'error': error,})
        else:
            error=1
            print('Formulario Erroneo')
            return render(request, 'main/colaborador_form.html', {'form': form, 'error': error})





def filtrar_candidatos(request):
    print(request)
    if request.method == 'GET':
        mesaa = Mesa.objects.filter(pk=int(request.GET.get('centro')))
        for m in mesaa:
            m1 = m.centro.municipio.pk
            d1 = m.centro.departamento.id

        centro = int(request.GET.get('centro'))
        presidentes = Candidato.objects.filter(tipo=1).order_by('tipo','partido__orden')
        diputados = Candidato.objects.filter(tipo=3,departamento=d1).order_by('tipo','partido__orden','casilla')
        candidatos = Candidato.objects.filter(tipo=2,municipio=m1).order_by('tipo','partido__orden')

        #c1 = presidencial.union(candidatos)
        #c2 = c1.union(diputados)

        #cfinal = c2.all().values('pk','tipo','partido','tipo__nombre', 'partido__nombre','nombre','foto').order_by('tipo','partido')
        #print(cfinal.query)
        centros_votacion = Escrutinio.objects.values('mesa').distinct()
        centros = Mesa.objects.exclude(pk__in=centros_votacion)
        # error = -1
        # #tipo_ = int(request.GET.get('tipo'))
        # departamento = int(request.GET.get('departamento'))
        # print(departamento)
        # #Presidencial
        #
        # if departamento == 19 and error == -1:
        #     print('Nivel Nacional')
        #     escrutinio = Escrutinio.objects.filter(tipo = tipo_, tipo_escrutinio = 'C').values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('valor')
        #     municipio = 0
        #     municipios = -1
        # #Sí es Departamental
        # if departamento > 0 and  departamento <= 71 and error == -1 :
        #     print('Departamental')
        #     escrutinio = Escrutinio.objects.filter(candidato__tipo = 3).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('num_votos')
        #     presidencial = Escrutinio.objects.filter(candidato__tipo = 1).values('candidato__nombre','candidato__foto').annotate(num_votos=Sum('valor')).order_by('num_votos')
        #
        #     print(escrutinio.query)
        #     municipios = Municipio.objects.all()
        #
        #
        # departamentos = Departamento.objects.all()
        # tipos = Tipo.objects.all()

        #identidad_ = request.GET.get('identidad')

        return render(request, "main/escrutinio_form.html",{ 'centro': centro,'alcaldes': candidatos,'presidentes':presidentes,'diputados': diputados, 'centros': centros},)



def votos_general(request):
    #if int(grado) <= 4:
    #    clases = Clase.objects.filter(grado__id = grado)
    template_path = 'estadisticas/votosg.html'
    today = timezone.now()

    context = None
    # context = {
    #     'alum': alumnos,
    #     'today': today,
    #     'request': request,
    #     'clases': clases,
    #     'grado': grado,
    #     'grado_': grado_
    # }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename"alumnos_todos.pdf"'
    template= get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(
       html, dest=response, link_callback=link_callback)
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

class ReporteEscrutiniosExcel(TemplateView):

    #Usamos el método get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        #print(resumendata.objects.all())
        resumen = resumendata.objects.filter(candidato_tipo ='Presidencial').order_by('centro_id','mesa_id','candidato_id')
        #print(resumen.query)
        #escrutinio = Escrutinio.objects.filter(candidato__tipo = 1).values('mesa__centro__nombre','mesa_id','mesa__numero','candidato__nombre').annotate(num_votos=Sum('valor')).order_by('mesa__numero','candidato__nombre')
        candidatos = Candidato.objects.filter(tipo = 1).order_by('partido')
        #escrutinio = Escrutinio.objects.raw('Select mesa_id as id ,candidato_id,mc.nombre , '+
        #                                        'sum(valor) from main_escrutinio me inner join main_candidato mc on me.candidato_id = mc.id where mc.tipo_id = 1 group by  mesa_id , candidato_id , mc.nombre, mc.casilla order by  mc.casilla')
        #.annotate(
        #    votos = sum()
        #)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE VOTOS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')

        Abc = ['A','B','C','D','E','F','G','H','I','J']
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'Departamento'
        ws['C3'] = 'Municipio'
        ws['D3'] = 'Centro'
        ws['E3'] = 'Mesa'
        columna = 6
        for c in candidatos:
            w = Abc[columna] + '3'
            #print(w)
            ws[w] = c.nombre
            columna += 1
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        #print(escrutinio)
        cont2 =7
        for esc in resumen:
            if cont == 4:
                mesa = esc.mesa
            #print(mesa,'-.-.-.-.-.-',esc.mesa ,'   /// ------ ', cont)

            if mesa == esc.mesa:
                for c in candidatos:
                    ws.cell(row=cont,column=2).value = esc.departamento
                    ws.cell(row=cont,column=3).value = esc.municipio
                    ws.cell(row=cont,column=4).value = esc.centro
                    ws.cell(row=cont,column=5).value = esc.mesa
                    ws.cell(row=cont,column=6).value = esc.candidato
                #if mesa == esc.mesa:
                #if c.nombre == esc.candidato:
                    cont2 = cont2 + 1
                    ws.cell(row=cont,column=cont2).value = esc.votos
                cont = cont + 1
            else:
                cont2 = 7

                print('diferente')
                #mesa = 0
            mesa = esc.mesa
            #print(mesa,'-.-.-.-.-.-',esc.mesa ,'   /// ------ ', cont)
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def dashboard_with_pivot(request):
    generar_estadistica()
    return render(request, 'estadisticas/dashboard_with_pivot.html', {})


def pivot_data(request):
    dataset = resumendata.objects.all()
    data = serializers.serialize('json', dataset)
    return JsonResponse(data, safe=False)

def generar_estadistica():
    escrutinio = Escrutinio.objects.all()
    #filter(candidato__tipo = 1).values('centro__departamento__pk','centro__municipio__pk','mesa__centro__pk','mesa__centro__nombre','mesa_id','mesa__numero','candidato__pk','candidato__nombre').annotate(num_votos=Sum('valor')).order_by('mesa__numero','candidato__nombre')
    for e in escrutinio:
        try:
            obj1, created = resumendata.objects.update_or_create(
                                mesa_id = e.mesa.pk,
                                mesa = e.mesa.numero,
                                partido = e.candidato.partido.nombre,
                                candidato_id = e.candidato.pk,
                                candidato = e.candidato.nombre,
                                defaults = {
                                    'departamento_id': e.mesa.centro.departamento.pk,
                                    'departamento': e.mesa.centro.departamento.nombre,
                                    'municipio_id' : e.mesa.centro.municipio.pk,
                                    'municipio': e.mesa.centro.municipio.nombre,
                                    'centro_id' : e.mesa.centro.pk,
                                    'centro': e.mesa.centro.nombre,
                                    'candidato_tipo': e.candidato.tipo.nombre,
                                    'votos': e.valor
                                    }
                                    )
            print('Todo correcto')
        except Exception as e:
            print('Error en votos /*/*/*/*/*/' , e)
